from openpyxl import load_workbook # библиотека для работы с таблицей excel
import numpy as np # библиотека для работы с матрицей

matrix = [] # исходная матрица смежности

def FindWay(current_matrix,v1,v2,count): # находит расстояние между вершинами
    way = 1 # степень матрицы
    while (current_matrix[v1-1][v2-1] == 0): # проход цикла по поиску расстояния (первое
                                            # ненулевое значение ячейки)
        current_matrix = np.dot(current_matrix,matrix) # возведение матрицы в след. степень
        way += 1
        if (way == count): # случай когда вершины не имеют связей (граф не связанный)
            return -1
    return way

def FindCount(current_matrix,v1,v2,way): # нахождение количества маршрутов длины way
    for i in range (1,way): # возведение матрицы в степень way
        current_matrix = np.dot(current_matrix, matrix)
    return current_matrix[v1-1][v2-1] # пересечение вершин в матрице -
                            # количество маршрутов длины, равной степени матрицы


def main():
    current_matrix = matrix
    count = len(current_matrix) # количестов вершин
    while (True):
        try:
            choice = int(input("Выберите действие:\n" # выбор действия с проверкой (все остальные
                               "1 - найти расстояние между вершинами\n" # пользовательские вводы также проверяются)
                               "2 - найти количество маршрутов определённой длины между вершинами\n"
                               "Выбор: "))
        except(ValueError):
            print ("Вы ввели не число")
            continue
        if (choice != 1) and (choice != 2):
            print("Введите 1 или 2 !!!")
            continue
        print("")
        if (choice == 1):
            try:
                v1 = int(input("Первая вершина: ")) # считывание номера первой вершины
                if (v1 > count) or (v1 < 1):
                    print("Такой вершины не существует!!!")
                    continue
                v2 = int (input("Вторая вершина: ")) # считывание номера второй вершины
                if (v2 > count) or (v2 < 1):
                    print("Такой вершины не существует!!!")
                    continue
            except(ValueError):
                print("Вы ввели не число")
                continue
            way = FindWay(current_matrix,v1,v2,count) # way - расстояние между v1 и v2
            if (way == -1):
                print(f"Вершины v{v1} и v{v2} не связаны")
            else:
                print("Расстояние: "+str(way))
            print("")

        if (choice == 2):
            try:
                v1 = int(input("Первая вершина: ")) # считывание номера первой вершины
                if (v1 > count) or (v1 < 1):
                    print("Такой вершины не существует!!!")
                    continue
                v2 = int(input("Вторая вершина: ")) # считывание номера второй вершины
                if (v2 > count) or (v2 < 1):
                    print("Такой вершины не существует!!!")
                    continue
                way = int(input("Длина маршрута: ")) # считывание длины маршрута
                if (way < 1):
                    print("Длина маршрута должна быть больше 0 !!!")
                    continue
            except(ValueError):
                print("Вы ввели не число")
                continue
            number = FindCount(current_matrix,v1,v2,way) # кол-во маршрутов длины way
            print(f"Количество маршрутов длины {way} между вершинами "
                  f"v{v1} и v{v2} равно {number}")

        print("")
        while (True):
            try:
                close = int(input("Выберите действие:\n" # диалог о прекращении программы
                                   "1 - ещё раз\n"
                                   "2 - закрыть\n"
                                   "Выбор: "))
            except(ValueError):
                print("Вы ввели не число")
                continue

            if (close != 1) and (close != 2):
                print("Введите 1 или 2 !!!")
                continue
            break
        if (close == 2):
            break
        print("")



def ReadFile(count): # чтение файла excel
    global matrix
    row = [] # строка матрицы
    try:
        wb = load_workbook("Matrix.xlsx") # загрузка таблицы
    except():
        print("Не удалось загрузить файл excel")
        return False
    sheet = wb['Matrix'] # выбор листа
    if (sheet.max_row != sheet.max_column): # проверка на квадратность матрицы
        print("Матрица должна быть квадратной!!!")
        return False

    if (sheet.max_row != count+1): # проверка на соответствие размера матрицы указанной длине
        print("Кол-во вершин в матрице не совпадает с указанным!!!\n"
              "Первые столбец и строка зарезервированы для вершин!!!")
        return False

    for i in range (count): # заполнение матрицы
        for j in range (count): # заполнение строки
            value = sheet.cell(row=2 + i, column=2 + j).value # доступ к значению ячейки
            if (value != None) and (type(value) == int ) and (value > -1):
                row.append(sheet.cell(row = 2 + i, column = 2 + j).value) # добавление элемента в строку
            else:
                print("В таблице имеется пустая/не целочисленная/отрицательная ячейка!!!")
                return False
        matrix.append(row) # добавление строки в матрицу
        row = []
    return True

while (True):
    try:
        vertex_count = int(input("Введите количестов вершин графа: "))
    except(ValueError):
        print("Вы ввели не число")
        continue
    if (vertex_count < 1):
        print("Количестов строк не может быть отрицательным числом!!!")
        continue
    break

if (ReadFile(vertex_count)):
    main()



